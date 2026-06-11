"""Carrega contatos de planilhas CSV e XLSX."""

from __future__ import annotations
import csv
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

PHONE_COLUMN_ALIASES = {"telefone", "celular", "whatsapp", "numero", "phone", "number"}
NAME_COLUMN_ALIASES = {"nome", "contato", "cliente", "name"}
MESSAGE_COLUMN_ALIASES = {"mensagem", "message", "texto", "recado"}
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}

class ContactSheetError(Exception):
    """Erro de leitura ou estrutura da planilha de contatos."""

@dataclass(frozen=True)
class ContactRow:
    row_number: int
    phone_number: str
    name: str
    message: str
    values: dict[str, str]

@dataclass(frozen=True)
class SheetRecords:
    headers: list[str]
    rows: list[dict[str, str]]

def load_contact_rows(
    file_path: str,
    phone_column: str = "",
    name_column: str = "",
    message_column: str = "",
) -> list[ContactRow]:
    path = resolve_sheet_path(file_path)
    records = read_sheet_records(path)
    columns = resolve_contact_columns(records.headers, phone_column, name_column, message_column)
    return [
        build_contact_row(row_number, row, columns)
        for row_number, row in enumerate(records.rows, start=2)
        if row_has_content(row)
    ]

def resolve_sheet_path(file_path: str) -> Path:
    path = Path(file_path).expanduser()

    if not path.exists():
        raise ContactSheetError(f"Arquivo nao encontrado: {path}")

    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ContactSheetError("Use uma planilha .csv, .xlsx ou .xlsm.")

    return path

def read_sheet_records(path: Path) -> SheetRecords:
    if path.suffix.lower() == ".csv":
        return read_csv_records(path)

    return read_excel_records(path)

def read_csv_records(path: Path) -> SheetRecords:
    with path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = clean_headers(reader.fieldnames or [])
        rows = [clean_record(row) for row in reader]

    ensure_headers(headers)
    return SheetRecords(headers=headers, rows=rows)

def load_openpyxl():
    try:
        import openpyxl
        return openpyxl.load_workbook
    except ImportError:
        raise ContactSheetError("A biblioteca openpyxl e necessaria para ler arquivos Excel. Instale com 'pip install openpyxl'.")
    
def read_excel_records(path: Path) -> SheetRecords:
    load_workbook = load_openpyxl()
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = clean_headers(next(rows, ()))
        records = rows_to_records(headers, rows)
    finally:
        workbook.close()

    ensure_headers(headers)
    return SheetRecords(headers=headers, rows=records)

def rows_to_records(headers: list[str], rows: Any) -> list[dict[str, str]]:
    return [
        dict(zip(headers, clean_cells(row), strict=False))
        for row in rows
    ]

def resolve_contact_columns(
    headers: list[str],
    phone_column: str,
    name_column: str,
    message_column: str,
) -> dict[str, str]:
    return {
        "phone": resolve_required_column(headers, phone_column, PHONE_COLUMN_ALIASES, "telefone"),
        "name": resolve_optional_column(headers, name_column, NAME_COLUMN_ALIASES),
        "message": resolve_optional_column(headers, message_column, MESSAGE_COLUMN_ALIASES),
    }

def resolve_required_column(
    headers: list[str],
    selected_column: str,
    aliases: set[str],
    column_description: str,
) -> str:
    column = resolve_column(headers, selected_column, aliases)

    if column:
        return column

    raise ContactSheetError(f"Nao encontrei a coluna de {column_description}.")

def resolve_optional_column(headers: list[str], selected_column: str, aliases: set[str]) -> str:
    return resolve_column(headers, selected_column, aliases)

def resolve_column(headers: list[str], selected_column: str, aliases: set[str]) -> str:
    normalized_headers = {normalize_text(header): header for header in headers if header}
    selected_key = normalize_text(selected_column)

    if selected_key:
        return find_selected_column(normalized_headers, selected_key, selected_column)

    for alias in aliases:
        if alias in normalized_headers:
            return normalized_headers[alias]

    return ""

def find_selected_column(
    normalized_headers: dict[str, str],
    selected_key: str,
    selected_column: str,
) -> str:
    if selected_key in normalized_headers:
        return normalized_headers[selected_key]

    raise ContactSheetError(f"Coluna nao encontrada: {selected_column}")

def build_contact_row(
    row_number: int,
    row: dict[str, str],
    columns: dict[str, str],
) -> ContactRow:
    return ContactRow(
        row_number=row_number,
        phone_number=row.get(columns["phone"], ""),
        name=row.get(columns["name"], "") if columns["name"] else "",
        message=row.get(columns["message"], "") if columns["message"] else "",
        values=row,
    )

def clean_headers(headers: Any) -> list[str]:
    return [cell_to_text(header) for header in headers]

def clean_record(row: dict[str, Any]) -> dict[str, str]:
    return {cell_to_text(key): cell_to_text(value) for key, value in row.items()}

def clean_cells(row: Any) -> list[str]:
    return [cell_to_text(cell) for cell in row]

def cell_to_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()

def ensure_headers(headers: list[str]) -> None:
    if not any(headers):
        raise ContactSheetError("A planilha precisa ter cabecalho na primeira linha.")

def row_has_content(row: dict[str, str]) -> bool:
    return any(value.strip() for value in row.values())

def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_text.lower().split())